"""
Generate PostgreSQL SQL scripts from the PitchBook data dictionary Excel file.

Reads table/column definitions from each sheet and produces:
  - sql/{TableName}.sql  (CREATE TABLE + \COPY for each table)
  - justfile             (parallel-loading recipes via xargs -P)
"""

from __future__ import annotations

import os
import re
from dataclasses import dataclass
from typing import Optional

import openpyxl


EXCEL_FILE = "docs/Brown_University_Academic_Feed_w__Public_Financials_v1_1.xlsx"
OUTPUT_DIR = "sql"
SCHEMA = "main"
DATA_DIR = "data"
SKIP_SHEETS = {"TrackingChanges", "Summary"}

PRIMARY_TABLES = [
    "Company",
    "Deal",
    "Investor",
    "Fund",
    "LimitedPartner",
    "ServiceProvider",
    "Person",
]

TYPE_MAP = {
    "DATE": "date",
    "TEXT": "text",
    "DECIMAL": "float8",
    "INTEGER": "integer",
    "LONG": "bigint",
}


@dataclass(frozen=True)
class Column:
    name: str
    data_type: str
    comment: str


@dataclass(frozen=True)
class Table:
    name: str
    columns: tuple[Column, ...]
    is_primary: bool


def map_type(raw_type: str) -> str:
    """Convert a PitchBook data type string to a PostgreSQL type."""
    if not raw_type:
        return "text"
    raw = raw_type.strip()
    m = re.match(r"TEXT\((\d+)\)", raw, re.IGNORECASE)
    if m:
        return "text"
    upper = raw.upper()
    return TYPE_MAP.get(upper, "text")


def truncate_comment(comment: str, max_len: int = 80) -> str:
    """Truncate a comment string to max_len characters."""
    if not comment:
        return ""
    s = comment.strip().replace("'", "''")
    return s[:max_len] if len(s) > max_len else s


def parse_sheet(ws, sheet_name: str) -> Optional[Table]:
    """Parse a single worksheet into a Table definition."""
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return None

    # Table name comes from cell B1 (row 0, col 1)
    table_name = rows[0][1]
    if not table_name:
        return None
    table_name = str(table_name).strip()

    is_primary = table_name in PRIMARY_TABLES

    # Column definitions start at row 6 (index 5)
    columns = tuple(
        Column(
            name=str(row[1]).strip(),
            data_type=map_type(str(row[2]) if row[2] else ""),
            comment=truncate_comment(str(row[3]) if row[3] else ""),
        )
        for row in rows[5:]
        if row[1] is not None
    )

    return Table(name=table_name, columns=columns, is_primary=is_primary)


def parse_workbook(path: str) -> list[Table]:
    """Parse all table-definition sheets from the Excel workbook."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    tables = []
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        table = parse_sheet(wb[sheet_name], sheet_name)
        if table:
            tables.append(table)
    wb.close()
    return tables


def find_primary_key(table: Table) -> Optional[str]:
    """Find the primary key column for a primary entity table.

    Convention: the PK column is named {TableName}ID.
    """
    if not table.is_primary:
        return None
    pk_name = f"{table.name}ID"
    return pk_name if any(c.name == pk_name for c in table.columns) else None


def generate_table_sql(table: Table) -> str:
    """Generate the full SQL script for a single table."""
    qualified = f"{SCHEMA}.{table.name}"
    pk_col = find_primary_key(table)

    lines = [
        f"-- Auto-generated from data dictionary for table: {table.name}",
        f"-- Columns: {len(table.columns)}",
        "",
        f"DROP TABLE IF EXISTS {qualified} CASCADE;",
        "",
        f"CREATE TABLE {qualified} (",
    ]

    col_defs = []
    num_cols = len(table.columns)
    for i, col in enumerate(table.columns):
        pg_type = col.data_type
        constraint = " PRIMARY KEY" if (col.name == pk_col) else ""
        comma = "," if i < num_cols - 1 else ""
        comment_str = f"  -- {col.comment}" if col.comment else ""
        col_defs.append(f"    \"{col.name}\" {pg_type}{constraint}{comma}{comment_str}")

    lines.append("\n".join(col_defs))
    lines.append(");")
    lines.append("")

    # Add column comments via COMMENT ON for documentation
    for col in table.columns:
        if col.comment:
            lines.append(
                f"COMMENT ON COLUMN {qualified}.\"{col.name}\" IS '{col.comment}';"
            )

    if any(c.comment for c in table.columns):
        lines.append("")

    # Indexes for relation tables
    if not table.is_primary and table.columns:
        first_col = table.columns[0].name
        idx_name = f"idx_{table.name}_{first_col}".lower()
        lines.append(f"CREATE INDEX {idx_name} ON {qualified} (\"{first_col}\");")
        # Index on RowID if present
        if any(c.name == "RowID" for c in table.columns):
            lines.append(
                f"CREATE INDEX idx_{table.name.lower()}_rowid ON {qualified} (\"RowID\");"
            )
        lines.append("")

    # COPY command
    lines.append(
        f"\\COPY {qualified} FROM '{DATA_DIR}/{table.name}.csv' "
        f"WITH (FORMAT csv, HEADER true, ENCODING 'UTF8', NULL '');"
    )
    lines.append("")

    return "\n".join(lines)


def main():
    tables = parse_workbook(EXCEL_FILE)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    for table in tables:
        sql = generate_table_sql(table)
        path = os.path.join(OUTPUT_DIR, f"{table.name}.sql")
        with open(path, "w") as f:
            f.write(sql)

    # Clean up old load_all.sql if it exists
    load_all_path = os.path.join(OUTPUT_DIR, "load_all.sql")
    if os.path.exists(load_all_path):
        os.remove(load_all_path)

    total_cols = sum(len(t.columns) for t in tables)
    print(f"Generated {len(tables)} table scripts in {OUTPUT_DIR}/")
    print(f"Total columns: {total_cols}")
    print(f"Primary tables: {[t.name for t in tables if t.is_primary]}")
    print(f"Relation tables: {len([t for t in tables if not t.is_primary])}")


if __name__ == "__main__":
    main()
